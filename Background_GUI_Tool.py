from tkinter import ttk
from datetime import datetime
import openpyxl
from tkinter import *
from HtmlDataExtraction.ComponentNetPin_GUI import NetPin_Gui
from HtmlDataExtraction.Netlist_GUI import Netlist_Gui
from HtmlDataExtraction.BOM_GUI import BOM_Gui
from HtmlDataExtraction.LenWidLayer_GUI import LenWidLaper_Gui
from DataConcatenation.DataConcat_GUI import DataConcat_Gui
from RefDesignator.RefDesignator_GUI import RefDesignator_Gui
from tkinter import messagebox
from HtmlDataExtraction.LayerStackup_GUI import LyrStack_Gui
from RefDesignator.LayerTable import CreateChlTable
from vlookup.vlookupMem_GUI import vlookup_gui
from RefDesignator.DdrLength import DdrLength_val_insertTOexcel

class Background_GUI:
    def __init__(self):
        self.BackGui_root = Tk()
        self.BackGui_root.title("Development Tool")
        self.BackGui_root.protocol("WM_DELETE_WINDOW",self.BackGui_root.withdraw)

        self.BackGui_root.withdraw()

        #---------------------Create the Notebook widget------------------------#
        self.notebook = ttk.Notebook(self.BackGui_root)

        # Create the first tab
        self.tab1 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab1, text='HTML Data Extraction')
        self.notebook.grid(columnspan=2)

        # Create the second tab
        self.tab2 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab2, text='Data Concatenation')
        self.notebook.grid()

        # Create the third tab
        self.tab3 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab3, text='REFDES')
        self.notebook.grid()

        # Create the fourth tab
        self.tab4 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab4, text='Vlookup Table')
        self.notebook.grid()

        #----------------Create New Report Workbook---------------#
        Date_Time=datetime.now().strftime('%Y-%m-%d %H%M%S')
        # y =str("Design Connectivity Checker_"+Date_Time)+'.'+'xlsx'
        # self.report_folder_path =  '/Users/yeamboonkai/Desktop/AMD_Project/Input_Files/DesignConnectivityChecker_Tool/Report_Output'
        self.report_folder_path =  'C:/DesignConnectivityChecker_Tool/Report_Output'
        self.fileName =self.report_folder_path + "//"+ str(Date_Time)+'.'+'xlsx'

        # Create a new workbook
        self.workbook = openpyxl.Workbook()

        # Remove the default Sheet
        default_sheet = self.workbook['Sheet']
        self.workbook.remove(default_sheet)

        # Create a new sheet
        self.sheet1 = self.workbook.create_sheet('Pin_Net')
        self.sheet2 = self.workbook.create_sheet('Netlist')
        self.sheet3 = self. workbook.create_sheet('BOM')
        self.sheet4 = self.workbook.create_sheet('NetWidth')
        self.sheet5 = self.workbook.create_sheet('Layer_Stackup')

        #---------------------------------------------------------#
        self.NetPin_Gui_obj = NetPin_Gui(self.tab1,self.sheet1)
        self.Netlist_Gui_obj = Netlist_Gui(self.tab1,self.sheet2)
        self.BOM_Gui_obj = BOM_Gui(self.tab1,self.sheet3)
        self.LenWidLayer_Gui_obj = LenWidLaper_Gui(self.tab1,self.sheet4)
        self.LayerStackup_Gui_obj = LyrStack_Gui(self.tab1,self.sheet5)
        self.DataConcat_Gui_obj = DataConcat_Gui(self.tab2,self.fileName)
        self.RefDesignator_Gui_obj = RefDesignator_Gui(self.tab3,self.fileName)
        self.vlookup_Gui_obj = vlookup_gui(self.tab4,self.fileName)

        #-------------------Create the export button--------------------------#
        self.export_count =  0
        export_button = Button(self.BackGui_root,text="Generate Report", command=self.Generate_Report)
        export_button.grid(row=4,column=0)

        #-------------------Hide, Show,Exit Button-------------------------------#
        presshide = Button(self.BackGui_root, text="Close Development Tool",command=self.hidebackground)
        presshide.grid(row=4,column=1)

        #-------------------Show report Generate Progress Status on Popup Message Box --------------------------#
        # Create a status bar label
        self.status_var = StringVar()
        self.status_label = ttk.Label(self.tab1, textvariable=self.status_var, anchor=W)
        self.status_label.grid(row=5, column=0,sticky="n")

        # Create a progressbar
        self.progressbar = ttk.Progressbar(self.tab1, length=1000, mode='determinate',orient=HORIZONTAL)
        self.progressbar.grid(row=6, column=0, pady=10)

    def progress_status(self,progress,status):
        #--------------Update progress status------------------#
        self.BackGui_root.update_idletasks()  # Update the GUI
        self.status_var.set(f"Generating {status}...{progress}%")
        self.progress['value'] = progress


    def Generate_Report(self):
        global progress  # Declare progress as a global variable
        self.popup = Toplevel(self.BackGui_root)
        self.popup.title("Progress")
        self.popup.geometry("1050x100")

        # Create a status bar label
        self.status_var = StringVar()
        self.status_label = ttk.Label(self.popup, textvariable=self.status_var, anchor=W)
        self.status_label.grid(row=0, column=0,sticky="n")

        # Create the progress bar
        self.progress = ttk.Progressbar(self.popup, orient="horizontal", length=1000, mode="determinate")
        self.progress.grid(row=1,column=0,pady=10,padx=20)

        # Start the long-running task
        self.popup.after(1, self.export_report)

        # # Handle the pop-up window closure
        # self.popup.protocol("WM_DELETE_WINDOW", lambda: messagebox.showinfo("Cannot Close", "Please wait for the task to finish."))

    def export_report(self):
        if self.export_count == 0:
            pass
        else:
            #Re-Create A New Report Workbook for 2nd times button click to avoid reuse existing sheetname#
            Date_Time = datetime.now().strftime('%Y-%m-%d %H%M%S')
            self.New_fileName = self.report_folder_path +"//"+ str(Date_Time)+'.'+'xlsx'

            # Create a new workbook
            self.workbook = openpyxl.Workbook()

            # Remove the default Sheet
            default_sheet = self.workbook['Sheet']
            self.workbook.remove(default_sheet)

            # Create a new sheet
            self.New_sheet1 = self.workbook.create_sheet('Pin_Net')
            self.New_sheet2 = self.workbook.create_sheet('Netlist')
            self.New_sheet3 = self. workbook.create_sheet('BOM')
            self.New_sheet4 = self.workbook.create_sheet('NetWidth')
            self.New_sheet5 = self.workbook.create_sheet('Layer_Stackup')

            # Update object's attribute of the new file created:
            self.NetPin_Gui_obj.sheet = self.New_sheet1
            self.Netlist_Gui_obj.sheet =self.New_sheet2
            self.BOM_Gui_obj.sheet = self.New_sheet3
            self.LenWidLayer_Gui_obj.sheet = self.New_sheet4
            self.LayerStackup_Gui_obj.sheet = self.New_sheet5
            self.DataConcat_Gui_obj.fileName = self.New_fileName
            self.RefDesignator_Gui_obj.fileName = self.New_fileName
            self.vlookup_Gui_obj.fileName = self.New_fileName

        # print(self.export_count, "click count check")

        self.input_check = {}
        self.input_check["Netpin"] = self.NetPin_Gui_obj.NetPin_input_entry.get()
        self.input_check["Netlist"] = self.Netlist_Gui_obj.Netlist_input_entry.get()
        self.input_check["BOM"] = self.BOM_Gui_obj.BOM_input_entry.get()
        self.input_check["NetWidth"] = self.LenWidLayer_Gui_obj.LenWidLayer_input_entry.get()
        self.input_check["Layer Stackup"] = self.LayerStackup_Gui_obj.LyrStack_input_entry.get()
        if all(value for value in self.input_check.values()):

            self.progress_status(progress=10,status="Net Pin data")
            self.NetPin_Gui_obj.NetPinOutputData()
            self.progress_status(progress=15,status="Net List data")
            self.Netlist_Gui_obj.NetlistOutputData()
            self.progress_status(progress=25,status="BOM data")
            self.BOM_Gui_obj.BomOutputData()
            self.progress_status(progress=30,status="Length Width Layer data")
            self.LenWidLayer_Gui_obj.LenWidLayerOutputData()
            self.progress_status(progress=40,status="Layer Stackup data")
            self.LayerStackup_Gui_obj.LyrStackupOutputData()
            

            if self.export_count == 0:
                # Save the workbook
                self.workbook.save(self.fileName)
            else:
                # Save the workbook
                self.workbook.save(self.New_fileName)

            self.progress_status(progress=45,status="Data Concatenation")
            self.DataConcat_Gui_obj.DataConcat_Execute()
            self.progress_status(progress=48,status="Copy Interface Memory Source data")
            self.RefDesignator_Gui_obj.copy_InterfaceMmy_data_to_dst()
            

            #------------------------------Create Memory Stackup Table------------------------------#
            self.Ch_Name =  ["Channel","ChA","ChB","ChC","ChD","ChE","CHF","ChG","ChH","ChI","ChJ","ChK","ChL"]

            # Get Breakout DQ input
            self.BO_width_DQ_Output_list = []
            for bodq_out in self.RefDesignator_Gui_obj.RefDsn_BO_Chnl_width_DQ_entry_list:
                try:
                    if len(bodq_out.get()) == 0:
                        self.BO_width_DQ_Output_list.append(bodq_out.get())
                    else:
                        self.float_convert = float(bodq_out.get())
                        self.BO_width_DQ_Output_list.append(self.float_convert)
                except:
                     messagebox.showinfo("Popup!", "Please enter only a number in Breakout/Bus Channel DQ & DQS input Field.")
                     break
            self.BO_DQ = ["Breakout DQ"] + self.BO_width_DQ_Output_list

            # Get Breakout DQS input
            self.BO_width_DQS_Output_list = []
            for bodqs_out in self.RefDesignator_Gui_obj.RefDsn_BO_Chnl_width_DQS_entry_list:
                try:
                    if len(bodqs_out.get()) == 0:
                        self.BO_width_DQS_Output_list.append(bodqs_out.get())
                    else:
                        self.float_convert = float(bodqs_out.get())
                        self.BO_width_DQS_Output_list.append(self.float_convert)
                except:
                     messagebox.showinfo("Popup!", "Please enter only a number in Breakout/Bus Channel DQ & DQS input Field.")
                     break
            self.BO_DQS = ["Breakout DQS"] + self.BO_width_DQS_Output_list

            # Get Bus Channel DQ input
            self.Bus_width_DQ_Output_list = []
            for busdq_out in self.RefDesignator_Gui_obj.RefDsn_Bus_Chnl_width_DQ_entry_list:
                try:
                    if len(busdq_out.get()) == 0:
                        self.Bus_width_DQ_Output_list.append(busdq_out.get())
                    else:
                        self.float_convert = float(busdq_out.get())
                        self.Bus_width_DQ_Output_list.append(self.float_convert)
                except:
                     messagebox.showinfo("Popup!", "Please enter only a number in Breakout/Bus Channel DQ & DQS input Field.")
                     break
            self.Bus_DQ = ["Bus Channel DQ"] + self.Bus_width_DQ_Output_list

            # Get Bus Channel DQS input
            self.Bus_width_DQS_Output_list = []
            for busdqs_out in self.RefDesignator_Gui_obj.RefDsn_Bus_Chnl_width_DQS_entry_list:
                try:
                    if len(busdqs_out.get()) == 0:
                        self.Bus_width_DQS_Output_list.append(busdqs_out.get())
                    else:
                        self.float_convert = float(busdqs_out.get())
                        self.Bus_width_DQS_Output_list.append(self.float_convert)
                except:
                     messagebox.showinfo("Popup!", "Please enter only a number in Breakout/Bus Channel DQ & DQS input Field.")
                     break
            self.Bus_DQS = ["Bus Channel DQ"] + self.Bus_width_DQS_Output_list


            if self.export_count == 0:
                self.progress_status(progress=53,status="Stackup Table data")
                # Create  stackup on “MEMORY” tab
                CreateChlTable(self.fileName,"MEMORY",self.Ch_Name,"F")
                CreateChlTable(self.fileName,"MEMORY",self.BO_DQ,"H")
                CreateChlTable(self.fileName,"MEMORY",self.BO_DQS,"I")
                CreateChlTable(self.fileName,"MEMORY",self.Bus_DQ,"J")
                CreateChlTable(self.fileName,"MEMORY",self.Bus_DQS,"K")
                DdrLength_val_insertTOexcel(fileName= self.fileName,
                                    Value= int(self.RefDesignator_Gui_obj.DDR_Length_val_entry.get()),
                                    sheet_name= self.RefDesignator_Gui_obj.DdrLgth_drop_click.get(),
                                    col_insert= self.RefDesignator_Gui_obj.DDR_Length_col_insert.get(),
                                    row_insert= int(self.RefDesignator_Gui_obj.DDR_Length_row_insert.get()),
                                    header= self.RefDesignator_Gui_obj.DDR_Length_header.get())
                
            else:
                self.progress_status(progress=53,status="Stackup Table data")
                # Create  stackup on “MEMORY” tab
                CreateChlTable(self.New_fileName,"MEMORY",self.Ch_Name,"F")
                CreateChlTable(self.New_fileName,"MEMORY",self.BO_DQ,"H")
                CreateChlTable(self.New_fileName,"MEMORY",self.BO_DQS,"I")   
                CreateChlTable(self.New_fileName,"MEMORY",self.Bus_DQ,"J")
                CreateChlTable(self.New_fileName,"MEMORY",self.Bus_DQS,"K")
                DdrLength_val_insertTOexcel(fileName= self.New_fileName,
                                    Value= int(self.RefDesignator_Gui_obj.DDR_Length_val_entry.get()),
                                    sheet_name= self.RefDesignator_Gui_obj.DdrLgth_drop_click.get(),
                                    col_insert= self.RefDesignator_Gui_obj.DDR_Length_col_insert.get(),
                                    row_insert= int(self.RefDesignator_Gui_obj.DDR_Length_row_insert.get()),
                                    header= self.RefDesignator_Gui_obj.DDR_Length_header.get())
                
            
            self.progress_status(progress=56,status="Memory Cpu0 Layer Stackup vlookup")
            self.vlookup_Gui_obj.run_MemCpu0_LyrStackup_Tbl_vlookup()
            
            """
             The get() method of the Entry widget is returning a string 
             with leading or trailing whitespace characters. 
             
             In such cases, using .strip() is a good practice to remove any 
             leading or trailing whitespace characters from the user input 
             before performing the comparison with an empty string ("").
            """
            if self.RefDesignator_Gui_obj.CPU_Ref_input_entry_widgets[0].get().strip() != "":
                self.progress_status(progress=59,status="Memory Referance Concat")
                self.RefDesignator_Gui_obj.Run_RefDesign_Concat()
                self.progress_status(progress=62,status="Memory Cpu0 NetName vlookup")
                self.vlookup_Gui_obj.run_MemCpu0_vlookup_Netnm()
                self.progress_status(progress=66,status="Memory Cpu0 Routing Layer data")
                self.vlookup_Gui_obj.run_MemCpu0_vlookup_RoutLyr()
                self.progress_status(progress=69,status="Memory Cpu0 Total Length data")
                self.vlookup_Gui_obj.run_MemCpu0_vlookup_TtlLgth()
                self.progress_status(progress=72,status="Memory Cpu0 Route Per MBDG data")
                self.vlookup_Gui_obj.run_MemCpu0_vlookup_RoutePerMbdg()
                self.progress_status(progress=75,status="Memory Cpu0 Breakout Length data")
                self.vlookup_Gui_obj.run_MemCpu0_vlookup_BoLength()
                self.progress_status(progress=78,status="Memory Cpu0 Impedance data")
                self.vlookup_Gui_obj.run_MemCpu0_vlookup_Impedance()
                # self.progress_status(progress=79,status="Memory Cpu0 Channel and Subclass Name")
                # self.vlookup_Gui_obj.run_MemCpu0_ChlLyrTbl_vlookup()
                # self.progress_status(progress=79.5,status="Memory Cpu0 Getting Min and Max Breakout Length")
                # self.vlookup_Gui_obj.run_MemCpu0_BoMinMax_vlookup()

                

            if self.RefDesignator_Gui_obj.CPU_Ref_input_entry_widgets[1].get().strip() != "":
                self.progress_status(progress=80,status="Memory Referance Concat")
                self.RefDesignator_Gui_obj.Run_RefDesign_Concat()
                self.progress_status(progress=81,status="Memory Cpu1 NetName vlookup")
                self.vlookup_Gui_obj.run_MemCpu1_vlookup_Netnm()
                self.progress_status(progress=84,status="Memory Cpu1 Routing Layer data")
                self.vlookup_Gui_obj.run_MemCpu1_vlookup_RoutLyr()
                self.progress_status(progress=87,status="Memory Cpu1 Total Length data")
                self.vlookup_Gui_obj.run_MemCpu1_vlookup_TtlLgth()
                self.progress_status(progress=90,status="Memory Cpu1 Route Per MBDG data")
                self.vlookup_Gui_obj.run_MemCpu1_vlookup_RoutePerMbdg()
                self.progress_status(progress=93,status="Memory Cpu1 Breakout Length data")
                self.vlookup_Gui_obj.run_MemCpu1_vlookup_BoLength()
                self.progress_status(progress=96,status="Memory Cpu1s Impedance data")
                self.vlookup_Gui_obj.run_MemCpu1_vlookup_Impedance()

            self.progress_status(progress=100, status="Complete")
            self.status_var.set("Generation complete!")

            if self.export_count == 0:
                # Task completed
                messagebox.showinfo("Task Complete" ,"Report Generate Complete  " + self.fileName)
            else:
                # Task completed
                messagebox.showinfo("Task Complete" ,"Report Generate Complete  " + self.New_fileName)

            self.popup.destroy()  # Close the main window
                
        else:
            for key, val in self.input_check.items():
                if not val:
                    messagebox.showinfo("Popup!", "The "+ key + " file input cannot be empty\nPlease select " +key+ " html file")
        self.export_count = self.export_count +1


    def hidebackground(self):
        self.BackGui_root.withdraw()

