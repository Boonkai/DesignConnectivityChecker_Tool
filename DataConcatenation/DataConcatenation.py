import openpyxl
import Background_GUI_Tool
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

class DataConcatenation:
    def __init__(self, 
                 fileName,
                 col_1= None,
                 col_2=None,
                 sheet1_name=None,
                 sheet2_name=None,
                 col_insert_sheet_name= None,
                 col_insert=None,
                 col_insert_start_row = None,
                 concat_symbol=None,
                 header = None):
        self.fileName =fileName
        self.col_1 = col_1
        self.col_2 = col_2
        self.sheet1_name = sheet1_name
        self.sheet2_name = sheet2_name
        self.col_insert = col_insert.upper()
        self.col_insert_start_row  = col_insert_start_row  # start from the specific row to exclude the header row
        self.concat_symbol = concat_symbol
        self.header = header
        self.col_insert_sheet_name = col_insert_sheet_name

    def data_concat(self):
        # Load the Excel file
        # workbook = openpyxl.load_workbook("2023-04-13 171628.xlsx")
        workbook = openpyxl.load_workbook(self.fileName)

        # Select the worksheet
        df1 = workbook[self.sheet1_name]
        df2 = workbook[self.sheet2_name]
        df3 = workbook[self.col_insert_sheet_name]

        end_row = df1.max_row

        # Set the border for each cell
        self.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        
        # set the alignment of the cell to center
        self.center_alignment = Alignment(horizontal='center', vertical='center')

        if len(self.col_insert) >=2:
            col_header_insert = 0
            for c in self.col_insert:
                col_header_insert = col_header_insert * 26 + ord(c) - ord('A') + 1
            print(col_header_insert)
        else:
            col_header_insert= ord(self.col_insert) - ord('A') + 1


        if self.header != None:
            df3.cell(row=1, column=col_header_insert).value = self.header

            # Set Alignment and border for header cell
            df3.cell(row=1, column=col_header_insert).alignment =self.center_alignment
            df3.cell(row=1, column=col_header_insert).border = self.border
        else:
            pass

        if self.col_insert_start_row != None:
            pass
        else:
            self.col_insert_start_row = 2


        for row in range(self.col_insert_start_row, end_row+1):
            column_1_value = df1[self.col_1 + str(row)].value
            column_2_value = df2[self.col_2 + str(row)].value
            combined_value = str(column_1_value) + self.concat_symbol + str(column_2_value) # combine the two column values with a space in between

            if self.col_insert == None:
                print(combined_value)
            else:
                df3[self.col_insert + str(row)].value = combined_value # write the combined value to the third column (column C)    

            # Set Alignment and border for each cell
            df3[self.col_insert + str(row)].alignment = self.center_alignment
            df3[self.col_insert + str(row)].border = self.border


        workbook.save(self.fileName)