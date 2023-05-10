from bs4 import BeautifulSoup
import re
import Background_GUI_Tool
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

# Files declaration:
NetPin_file = "Component_Pin_Report.htm"
Netlist_file = "Netlist.htm"
BOM_file = "BOM.htm"
LenWidLayer_file = "Etch_Length_Width_Layer.htm"
LayerStack_file = "Layer_Stackup_report.htm"


class DataExtraction:
    # Provide required files for the first argument as follow:
    #1.NetPin_file OR
    #2.Netlist_file OR
    #3.BOM_file OR
    #4.TraceLengthWidthLayer_file

    # Provide column extract for the second argument:
    # e.g:
    # "col_index = [0,1,5]" --> represent 1st,2nd and 6th column data will be extract
    #NOTE: [0,1,5] is an index base where 1st index always start from Zero 

    def __init__(self,file,col_select):
        self.file = file
        self.col_select = col_select

        # Component_Pin_Report empty list:
        self.col_data_ComponentPin_REFDES = []
        self.col_data_ComponentPin_PinNumber = []
        self.col_data_ComponentPin_CompDeviceTyp = []
        self.col_data_ComponentPin_PinTyp= []
        self.col_data_ComponentPin_PinName = []
        self.col_data_ComponentPin_NetName = []

        # Netlist empty list:
        self.col_data_Netlist_NetName = []
        self.col_data_Netlist_NetPins = []
        self.col_data_Netlist_NetPins_1st_data = []
        self.col_data_Netlist_NetPins_2nd_data = [' ']

        # BOM Report empty list:
        self.col_data_BOM_SymName = []
        self.col_data_BOM_CompDeviceTyp = []
        self.col_data_BOM_CompValue = []
        self.col_data_BOM_CompTol = []
        self.col_data_BOM_CompClass = []
        self.col_data_BOM_REFDES = []

        # Trace Length by Layer and Width Report empty list:
        self.col_data_TraceLenWid_NetName = []
        self.col_data_TraceLenWid_LayerName = []
        self.col_data_TraceLenWid_TotalLength = []
        self.col_data_TraceLenWid_LineWidth = []
        self.col_data_TraceLenWid_LenATwidth = []

        #Layer Stackup
        self.Layer_num = 1
        self.col_layer_stackup_SubName = []
        self.col_layer_stackup_LayerName = ["Layer"]
        self.col_layer_stackup_Type = []
        self.col_layer_stackup_Material = []
        self.col_layer_stackup_Thickness = []


    def HTML_Data_Extract(self):
        # Open the HTML file
        with open(str(self.file), 'r') as html_file:
            # Read the file's contents
            html_content = html_file.read()

        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(html_content, 'html.parser')

        # Find the table element you want to extract data from
        table = soup.find("table")

        #Component Pin Report data extract:
        if NetPin_file in self.file:
            for row in table.find_all("tr"):
                columns = row.find_all("td")
                for i in self.col_select:
                    if i.get() == "NetPin_option1":
                        self.col_data_ComponentPin_REFDES.append(columns[0].text.strip())
                    if i.get() == "NetPin_option2":
                        # Using "Regular expression" to detect html mixed number (numbers that contain digits and non-digits)
                        # if mixed-number / empty ==> store in to list, else convert to integer and sotre to list.
                        if re.findall('[A-Z]+',columns[1].text.strip()):
                            self.col_data_ComponentPin_PinNumber.append(columns[1].text.strip())
                        elif len(columns[1].text.strip()) == 0:
                            self.col_data_ComponentPin_PinNumber.append(columns[1].text.strip())
                        else:
                            j = int(columns[1].text.strip())
                            self.col_data_ComponentPin_PinNumber.append(j)
                    if i.get() == "NetPin_option3":
                        self.col_data_ComponentPin_CompDeviceTyp.append(columns[2].text.strip())
                    if i.get() == "NetPin_option4":
                        self.col_data_ComponentPin_PinTyp.append(columns[3].text.strip())
                    if i.get() == "NetPin_option5":
                        self.col_data_ComponentPin_PinName.append(columns[4].text.strip())
                    if i.get() == "NetPin_option6":
                        self.col_data_ComponentPin_NetName.append(columns[5].text.strip())

        #Netlist Report data extract:
        if Netlist_file in self.file:
            for row in table.find_all("tr"):
                columns = row.find_all("td")
                for i in self.col_select:
                    if i.get() == "Netlist_option1":
                        self.col_data_Netlist_NetName.append(columns[0].text.strip())
                    if i.get() == "Netlist_option2":
                        self.col_data_Netlist_NetPins.append(columns[1].text.strip())
            
            # Separate out 2nd data afater spacing
            for index, separate in enumerate(self.col_data_Netlist_NetPins):
                if index == 0:
                    self.col_data_Netlist_NetPins_1st_data.append(separate)
                else:
                    atpos = separate.find(' ')
                    self.col_data_Netlist_NetPins_1st_data.append(separate[:atpos ])
                    self.col_data_Netlist_NetPins_2nd_data.append(separate[atpos: ])
                    # print(separate[atpos: ])

        #BOM Report data extract:
        if BOM_file in self.file:
            for row in table.find_all("tr"):
                columns = row.find_all("td")
                for i in self.col_select:
                    if i.get() == "BOM_option1":
                        self.col_data_BOM_SymName.append(columns[0].text.strip())
                    if i.get() == "BOM_option2":
                        self.col_data_BOM_CompDeviceTyp.append(columns[1].text.strip())
                    if i.get() == "BOM_option3":
                        self.col_data_BOM_CompValue.append(columns[2].text.strip())
                    if i.get() == "BOM_option4":
                        self.col_data_BOM_CompTol.append(columns[3].text.strip())
                    if i.get() == "BOM_option5":
                        self.col_data_BOM_CompClass.append(columns[4].text.strip())
                    if i.get() == "BOM_option6":
                        self.col_data_BOM_REFDES.append(columns[5].text.strip())

        # Trace Length by Layer and Width Report data extract:
        if LenWidLayer_file in self.file:
            for row in table.find_all("tr"):
                columns = row.find_all("td")
                for i in self.col_select:
                    if i.get() == "LenWidLayer_option1":
                        self.col_data_TraceLenWid_NetName.append(columns[0].text.strip())
                    if i.get() == "LenWidLayer_option2":
                        self.col_data_TraceLenWid_LayerName.append(columns[1].text.strip())
                    if i.get() == "LenWidLayer_option3":
                        # Using "Regular expression" to detect html mixed number (numbers that contain digits and non-digits)
                        # if mixed-number ==> store in to list, else convert to integer and sotre to list.
                        if re.findall('[A-Z]+',columns[2].text.strip()):
                            self.col_data_TraceLenWid_TotalLength.append(columns[2].text.strip())
                        else:
                            k = float(columns[2].text.strip())
                            self.col_data_TraceLenWid_TotalLength.append(k)
                    if i.get()== "LenWidLayer_option4":
                        if re.findall('[A-Z]+',columns[3].text.strip()):
                            self.col_data_TraceLenWid_LineWidth.append(columns[3].text.strip())
                        else:
                            l = float(columns[3].text.strip())
                            self.col_data_TraceLenWid_LineWidth.append(l)
                    if i.get() == "LenWidLayer_option5":
                        if re.findall('[A-Z]+',columns[4].text.strip()):
                            self.col_data_TraceLenWid_LenATwidth.append(columns[4].text.strip())
                        else:
                            m = float(columns[4].text.strip())
                            self.col_data_TraceLenWid_LenATwidth.append(m) 

        #layer Stackup Report data extract:
        if LayerStack_file in self.file:
            for row in table.find_all("tr"):
                columns = row.find_all("td")
                for i in self.col_select:
                    if i.get() == "LyrStack_option1":
                        self.col_layer_stackup_SubName.append(columns[0].text.strip())
                        if columns[0].text.strip() ==  "Subclass Name":
                            continue
                        # Detect layer name, store layer name into list if not empty else append empty
                        if len(columns[0].text.strip()) > 0 :
                            self.col_layer_stackup_LayerName.append("L" + str(self.Layer_num))
                            self.Layer_num += 1
                        else:
                            self.col_layer_stackup_LayerName.append(columns[0].text.strip())
                    if i.get() == "LyrStack_option2":
                        self.col_layer_stackup_Type.append(columns[1].text.strip())
                    if i.get() == "LyrStack_option3":
                        self.col_layer_stackup_Material.append(columns[2].text.strip())
                    if i.get() == "LyrStack_option4":
                        # Using "Regular expression" to detect html mixed number (numbers that contain digits and non-digits)
                        # if mixed-number ==> store in to list, else convert to integer and sotre to list.
                        if re.findall('[A-Z]+',columns[3].text.strip()):
                            self.col_layer_stackup_Thickness.append(columns[3].text.strip())
                        elif len(columns[3].text.strip()) == 0:
                            self.col_layer_stackup_Thickness.append(columns[3].text.strip())
                        else:
                            l = float(columns[3].text.strip())
                            self.col_layer_stackup_Thickness.append(l)

class ExportTOexcel:
    def __init__(self,object,sheet,column_num):
        self.row = 0
        self.object = object
        self.sheet = sheet
        self.column_num =column_num

    def WriteTOexcel(self):
        # print(len(self.object))
            # Set the border for each cell
        self.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
        for i in self.object:
            # Set Alignment and border for each cell
            self.sheet.cell(self.row +1,self.column_num +1).value = i
            self.sheet.cell(self.row +1,self.column_num +1).alignment = self.center_alignment
            self.sheet.cell(self.row +1,self.column_num +1).border = self.border
            self.row += 1
    