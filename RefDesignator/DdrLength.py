import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

class DdrLength_val_insertTOexcel:
    def __init__(self,
                 fileName,
                 sheet_name,
                 Value,
                 col_insert,
                 row_insert,
                 header):
        self.fileName = fileName
        self.sheet_name = sheet_name
        self.Value = Value
        self.col_insert = col_insert
        self.row_insert = row_insert
        self.header = header

        workbook = openpyxl.load_workbook(self.fileName)
        df1 = workbook[self.sheet_name]

        # Set the border for each cell
        self.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        
        # set the alignment of the cell to center
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        # Fill up color Style on header cell  
        self.header_fill = PatternFill(start_color="92d050", end_color="92d050", fill_type="solid")
        # Fill up color Style on column 
        self.col_fill = PatternFill(start_color="fce4d6", end_color="fce4d6", fill_type="solid")

        df1[self.col_insert + str(row_insert)].value = self.Value

        # Set Alignment and border for each cell
        df1[self.col_insert + str(row_insert)].alignment = self.center_alignment
        df1[self.col_insert + str(row_insert)].border = self.border
        # df1[self.col_insert + str(row_insert)].fill = self.col_fill

        # Set Color for header cell
        df1[self.col_insert + str(row_insert-1)].value = self.header
        df1[self.col_insert + str(row_insert-1)].fill = self.header_fill
        df1[self.col_insert + str(row_insert-1)].alignment = self.center_alignment
        df1[self.col_insert + str(row_insert-1)].border = self.border

        workbook.save(self.fileName)

# class vlookup_DdrLength:
#     def run_vloopup_DDRLength(self):
#         vlookup( filename = DdrLength_val_insertTOexcel().fileName,
#                 sheet1 = DdrLength_val_insertTOexcel().sheet_name,
#                 lookup_Tbl_output= DdrLength_val_insertTOexcel().Value,
#                 lookup_out_insert = DdrLength_val_insertTOexcel().col_insert, 
#                 header= DdrLength_val_insertTOexcel().header ,
#                 )