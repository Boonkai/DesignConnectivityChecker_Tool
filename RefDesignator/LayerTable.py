import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

class CreateChlTable:
    def __init__(self,
                 fileName,
                 sheet_name,
                 TableValue,
                 col_insert,
                 start_row_insert,
                 merge_range = None,
                 merge_header = None,
                 header = None):
        self.fileName = fileName
        self.sheet_name = sheet_name
        self.TableValue = TableValue
        self.col_insert = col_insert
        self.start_row_insert = start_row_insert
        self.merge_range = merge_range # Define the range of cells to merge
        self.merge_header = merge_header
        self.header = header

        workbook = openpyxl.load_workbook(self.fileName)
        df1 = workbook[self.sheet_name]

        # Set the border for each cell
        self.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        
        # set the alignment of the cell to center
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        # Fill up color Style on header cell  
        self.header_fill = PatternFill(start_color="92d050", end_color="92d050", fill_type="solid")
        # Fill up color Style on column 
        self.col_fill = PatternFill(start_color="fce4d6", end_color="fce4d6", fill_type="solid")

        for num, chl in enumerate(self.TableValue):
            df1[self.col_insert + str(num+self.start_row_insert)].value = chl

            # Set Alignment and border for each cell
            df1[self.col_insert + str(num+start_row_insert)].alignment = self.center_alignment
            df1[self.col_insert + str(num+start_row_insert)].border = self.border
            df1[self.col_insert + str(num+start_row_insert)].fill = self.col_fill

        # Set Color for header cell
        df1[self.col_insert + str(start_row_insert)].fill = self.header_fill

        try:
            if self.header is not None:
                # Set Color for header cell if header is True
                df1[self.col_insert + str(self.start_row_insert)].value = self.header
                df1[self.col_insert + str(start_row_insert)].fill = self.header_fill

            if self.merge_range is not None:
                # Merge the cells
                df1.merge_cells(merge_range)

                # Access the merged cell and set the alignment and color
                df1[self.merge_range.split(":")[0]].value = self.merge_header
                df1[self.merge_range.split(":")[0]].alignment = self.center_alignment
                df1[self.merge_range.split(":")[0]].border = self.border
                df1[self.merge_range.split(":")[0]].fill = PatternFill(start_color="c6e0b4", end_color="c6e0b4", fill_type="solid")
        except Exception as e:
            print("An error occurred:", str(e))
        
        workbook.save(self.fileName)

