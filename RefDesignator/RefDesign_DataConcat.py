import openpyxl
import Background_GUI_Tool
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

class RefDesig_Concat:
    def __init__(self, 
                 filename,
                 Ref_value = None,
                 concat_symbol=None,
                 Insert_sheet_name = None,
                 col_insert=None,
                 col_lookup = None,
                 col_insert_start_row = None,
                 header = None):
        
        self.filename =filename
        self.Ref_value = Ref_value
        self.col_lookup = col_lookup
        self.concat_symbol = concat_symbol
        self.Insert_sheet_name = Insert_sheet_name
        self.col_insert = col_insert.upper()
        self.col_insert_start_row  = int(col_insert_start_row)  # start from the specific row to exclude the header row
        self.header = header

    def data_concat(self):
        # Load the Excel file
        # workbook = openpyxl.load_workbook("2023-04-27 123135.xlsx")
        workbook = openpyxl.load_workbook(self.filename)

        # Select the worksheet
        df1 = workbook[self.Insert_sheet_name]

        end_row = df1.max_row

        # Set the border for each cell
        self.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        
        # set the alignment of the cell to center
        self.center_alignment = Alignment(horizontal='center', vertical='center')


        # print(self.col_insert)
        if len(self.col_insert) >=2:
            col_header_insert = 0
            for c in self.col_insert:
                col_header_insert = col_header_insert * 26 + ord(c) - ord('A') + 1
            print(col_header_insert)
        else:
            col_header_insert= ord(self.col_insert) - ord('A') + 1

        if self.header != None:
            df1.cell(self.col_insert_start_row-1, column=col_header_insert).value = self.header

            # Set Alignment and border for header cell
            df1.cell(self.col_insert_start_row-1, column=col_header_insert).alignment =self.center_alignment
            df1.cell(self.col_insert_start_row-1, column=col_header_insert).border = self.border

        else:
            pass

        for row in range(self.col_insert_start_row, end_row+1):
            lookup_value = df1[self.col_lookup + str(row)].value
            combined_value = str(self.Ref_value) + self.concat_symbol + str(lookup_value) # combine the two column values with a space in between
            df1[self.col_insert + str(row)].value = combined_value # write the combined value to the third column (column C)    
            
            # Set Alignment and border for each cell
            df1[self.col_insert + str(row)].alignment = self.center_alignment
            df1[self.col_insert + str(row)].border = self.border

        workbook.save(self.filename)
        # workbook.save("2023-04-27 123135.xlsx")