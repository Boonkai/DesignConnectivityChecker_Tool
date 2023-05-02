from tkinter import filedialog
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

def RefDesign_browse_file():
    # Select the source Excel file
    src_file = filedialog.askopenfilename(title="Select Source File", filetypes=[("Excel files", "*.xlsx")])
    if not src_file:
        return
    
    return src_file


def copy_tab(src_file,dst_file,src_sheetname):
    # print(src_file,dst_file)
    for i in src_sheetname:
        dst_ws = None
        if i.get() != str("0"):
            # print(i.get())
            src_wb = openpyxl.load_workbook(src_file)
            src_ws = src_wb[i.get()]

            dst_wb = openpyxl.load_workbook(dst_file)
            dst_ws = dst_wb.create_sheet(i.get())

            for row_insert in src_ws.iter_rows(values_only=True):
                dst_ws.append(row_insert)

            # Set the border for each cell
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))

            for row in dst_ws.iter_rows():
                for cell in row:

                    # Set Alignment and border for each cell
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    cell.alignment = center_alignment
                    cell.border = border

            dst_wb.save(dst_file)