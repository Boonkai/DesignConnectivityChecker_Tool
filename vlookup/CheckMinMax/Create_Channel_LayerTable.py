from tkinter import *
from vlookup.run_vlookup import vlookup
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

class ChannelLayerName_Table_Gui:
    def __init__(self,
                 rootframe,
                 first_sheet,
                 tabel_col,
                 result_col,
                 second_sheet,
                 insert_col_ChlNm,
                 insert_col_LyrNm,
                 insert_row,
                 ):
        self.rootframe = rootframe
        self.first_sheet = first_sheet
        self.tabel_col = tabel_col
        self.result_col = result_col
        self.second_sheet = second_sheet
        self.insert_col_ChlNm = insert_col_ChlNm
        self.insert_col_LyrNm = insert_col_LyrNm
        self.insert_row = insert_row

        self.lyrStack_fme = LabelFrame(self.rootframe, text="Channel & Layer Name Table",labelanchor="n")
        self.lyrStack_fme.grid(row=2,column=1)


        #------------------------First Sheet Gui-----------------------# 
        self.lyrStack_sheet_1 = LabelFrame(self.lyrStack_fme,text="First Sheet (Lookup Table)",labelanchor="n")
        self.lyrStack_sheet_1.grid(row=0,column=0,padx=20,pady=15)

        self.shee1_label = Label(self.lyrStack_sheet_1,text="Sheet:").grid(row=0,column=0,padx=3,pady=3)

        # self.Sheet1_drop_list = []
        self.Sheet1_options = [ "Pin_Net", 
                            "Netlist",
                            "BOM",
                            "NetWidth",
                            "Layer_Stackup",
                            "MEMORY",
                            "XGMI",
                            "PCIe",
                            "SATA",
                            "MISC",
                            "USB",
                            "CLK"]

        #Drop Down Boxes
        self.Sheet1_clicked = StringVar()
        self.Sheet1_clicked.set(self.Sheet1_options[int(self.first_sheet)])
        # self.Sheet1_drop_list.append(self.Sheet1_clicked)

        self.Sheet1_drop = OptionMenu(self.lyrStack_sheet_1,self.Sheet1_clicked,*self.Sheet1_options)
        self.Sheet1_drop.grid(row=0,column=1,padx=5,pady=5)

        # print(self.Sheet1_clicked.get())

        self.sheet1_lkupTbl_Col = Label(self.lyrStack_sheet_1,text="Lookup \nTable Col").grid(row=1,column=0,padx=5,pady=5,sticky="n")
        self.sheet1_lkupTbl_Col_Entry = Entry(self.lyrStack_sheet_1,borderwidth=3,width=10,fg="black",background="white")
        self.sheet1_lkupTbl_Col_Entry.grid(row=1,column=1,padx=3,pady=3)

        self.sheet1_lkupTbl_result_Col = Label(self.lyrStack_sheet_1,text="Result Col").grid(row=2,column=0,padx=5,pady=5,sticky="n")
        self.sheet1_lkupTbl_result_Entry = Entry(self.lyrStack_sheet_1,borderwidth=3,width=10,fg="black", background="white")
        self.sheet1_lkupTbl_result_Entry.grid(row=2,column=1,padx=3,pady=3)

        #------------------------Second Sheet Gui-----------------------# 
        self.lyrStack_sheet_2 = LabelFrame(self.lyrStack_fme,text="Second Sheet (Lookup value)",labelanchor="n")
        self.lyrStack_sheet_2.grid(row=0,column=1,padx=15,pady=15)

        self.shee2_label = Label(self.lyrStack_sheet_2,text="Sheet:").grid(row=0,column=0,padx=3,pady=3)

        # self.Sheet1_drop_list = []
        self.Sheet2_options = [ "Pin_Net", 
                            "Netlist",
                            "BOM",
                            "NetWidth",
                            "Layer_Stackup",
                            "MEMORY",
                            "XGMI",
                            "PCIe",
                            "SATA",
                            "MISC",
                            "USB",
                            "CLK"]

        #Drop Down Boxes
        self.Sheet2_clicked = StringVar()
        self.Sheet2_clicked.set(self.Sheet2_options[int(self.second_sheet)])
        # self.Sheet1_drop_list.append(self.Sheet1_clicked)

        self.Sheet2_drop = OptionMenu(self.lyrStack_sheet_2,self.Sheet2_clicked,*self.Sheet2_options)
        self.Sheet2_drop.grid(row=0,column=1,padx=5,pady=5)

        self.sheet2_lkup_Chl_result_Col = Label(self.lyrStack_sheet_2,text="ChannelName\nOutput Insert Col").grid(row=1,column=0,padx=5,pady=5,sticky="n")
        self.sheet2_result_Chl_insert_col_Entry = Entry(self.lyrStack_sheet_2,border=3,width=10,fg="black",background="white")
        self.sheet2_result_Chl_insert_col_Entry.grid(row=1,column=1,padx=3,pady=3)

        self.sheet2_lkup_Lyr_result_Col = Label(self.lyrStack_sheet_2,text="LayerName\nOutput Insert Col").grid(row=2,column=0,padx=5,pady=5,sticky="n")
        self.sheet2_result_Lyr_insert_col_Entry = Entry(self.lyrStack_sheet_2,border=3,width=10,fg="black",background="white")
        self.sheet2_result_Lyr_insert_col_Entry.grid(row=2,column=1,padx=3,pady=3)

        self.sheet2_lkup_result_row = Label(self.lyrStack_sheet_2,text="Result\nInsert row").grid(row=3,column=0,padx=5,pady=5,sticky="n")
        self.sheet2_result_insert_row_Entry = Entry(self.lyrStack_sheet_2,border=3,width=10,fg="black",background="white")
        self.sheet2_result_insert_row_Entry.grid(row=3,column=1,padx=3,pady=3)

        #Set Default input
        self.sheet1_lkupTbl_Col_Entry.insert(0,str(self.tabel_col))
        self.sheet1_lkupTbl_result_Entry.insert(0,str(self.result_col))
        self.sheet2_result_Chl_insert_col_Entry.insert(0,str(self.insert_col_ChlNm))
        self.sheet2_result_Lyr_insert_col_Entry.insert(0,str(self.insert_col_LyrNm))
        self.sheet2_result_insert_row_Entry.insert(0,str(self.insert_row))
    

    def run_CreateChllNmTbl(self,fileName):
        # Set the border for each cell                    
        self.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        # Set Alignment to each cell
        self.center_alignment = Alignment(horizontal='center', vertical='center')

        # Fill up color Style on header cell  
        self.header_fill = PatternFill(start_color="92d050", end_color="92d050", fill_type="solid")
        # Fill up color Style on column 
        self.col_fill = PatternFill(start_color="fce4d6", end_color="fce4d6", fill_type="solid")

        self.fileName = fileName
        # Load the Excel file containing the source and destination worksheets
        wb = openpyxl.load_workbook(self.fileName)

        # Select the source and destination worksheets
        destination_ws = wb[self.Sheet2_clicked.get()]

        # Channel Name list
        self.Ch_Name =  ["Channel","ChA","ChB","ChC","ChD","ChE","CHF","ChG","ChH","ChI","ChJ","ChK","ChL"]

        # Convert result insert row entry to integer
        self.row = int(self.sheet2_result_insert_row_Entry.get())

        # NOTE: Hard Code Channel Name row insert to =>"F"
        for i, name in enumerate(self.Ch_Name):
            if i == 0:
                destination_ws[self.sheet2_result_Chl_insert_col_Entry.get() + str(self.row+i)].value = name
                destination_ws[self.sheet2_result_Chl_insert_col_Entry.get() + str(self.row+i)].alignment = self.center_alignment
                destination_ws[self.sheet2_result_Chl_insert_col_Entry.get() + str(self.row+i)].border = self.border
                destination_ws[self.sheet2_result_Chl_insert_col_Entry.get() + str(self.row+i)].fill = self.header_fill
            else:
                destination_ws[self.sheet2_result_Chl_insert_col_Entry.get() + str(self.row+i)].value = name
                destination_ws[self.sheet2_result_Chl_insert_col_Entry.get() + str(self.row+i)].alignment = self.center_alignment
                destination_ws[self.sheet2_result_Chl_insert_col_Entry.get() + str(self.row+i)].border = self.border
                destination_ws[self.sheet2_result_Chl_insert_col_Entry.get() + str(self.row+i)].fill = self.col_fill
        wb.save(self.fileName)


    def run_ChlLyrTbl_vlookup(self,fileName):
        self.fileName = fileName
        # print(self.fileName)
        vlookup(filename= self.fileName,
                sheet1=self.Sheet1_clicked.get(),
                sheet2= self.Sheet2_clicked.get(),
                lookup_Tbl_column = self.sheet1_lkupTbl_Col_Entry.get(),
                lookup_Tbl_output = self.sheet1_lkupTbl_result_Entry.get(),
                lookup_out_insert = self.sheet2_result_Lyr_insert_col_Entry.get(),
                lookup_out_insert_row= self.sheet2_result_insert_row_Entry.get()).vlookup_Stackup_Table()


