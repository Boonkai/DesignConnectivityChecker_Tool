import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

class vlookup:
    def __init__(self,
                 filename,
                 sheet1,
                 sheet2,
                 lookup_out_insert = None,
                 lookup_Tbl_column = None,
                 lookup_Tbl_output = None,
                 lookup_val_col=None,
                 lookup_Tbl_LyrNm_out =None,
                 lookup_Tbl_TtlLgth_out = None,
                 lookup_lyrNm_out_insert =None,
                 lookup_TtlLgth_out_insert =None,
                 lookup_Tbl_col_LineWdt = None,
                 header_LyrNm =None,
                 header_TtlLgth =None,
                 header=None,
                 NetName_Col=None):
        
        
        self.filename = filename
        self.sheet1 = sheet1
        self.sheet2 = sheet2
        self.lookup_out_insert = lookup_out_insert
        self.lookup_Tbl_column = lookup_Tbl_column
        self.lookup_Tbl_col_LineWdt = lookup_Tbl_col_LineWdt
        self.lookup_Tbl_output = lookup_Tbl_output
        self.header = header
        self.NetName_Col = NetName_Col
        self.lookup_Tbl_LyrNm_out = lookup_Tbl_LyrNm_out
        self.lookup_Tbl_TtlLgth_out = lookup_Tbl_TtlLgth_out
        self.lookup_lyrNm_out_insert = lookup_lyrNm_out_insert
        self.lookup_TtlLgth_out_insert = lookup_TtlLgth_out_insert
        self.header_LyrNm = header_LyrNm
        self.header_TtlLgth = header_TtlLgth
        self.lookup_val_col = lookup_val_col

        # Open the Excel file
        self.workbook = openpyxl.load_workbook(self.filename)

        # Select the sheet to work with
        # sheet1 is table lookup whereas sheet2 is lookup value
        self.sheet1 = self.workbook[self.sheet1]
        self.sheet2 = self.workbook[self.sheet2]

        # Set the border for each cell                    
        self.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        # Set Alignment to each cell
        self.center_alignment = Alignment(horizontal='center', vertical='center')

    def column_label_to_number(self,label):
        total = 0
        for i, char in enumerate(reversed(label)):
            value = ord(char) - 64
            total += value * (26 ** i)
        return total
    
    def color_styling(self,headClr,ColClr):
        # Fill up color Style on header cell  
        self.header_fill = PatternFill(start_color=headClr, end_color=headClr, fill_type="solid")

        # Fill up color Style on column 
        self.col_fill = PatternFill(start_color=ColClr, end_color=ColClr, fill_type="solid")

        return self.header_fill, self.col_fill

    
    def vlookupTable(self):
        #Character to Int Conversion 
        self.lkvTbl_Col  = self.column_label_to_number(self.lookup_Tbl_column.upper())
        self.lkvTbl_Result_Col = self.column_label_to_number(self.lookup_Tbl_output.upper())
        self.lkv_out_insert = self.column_label_to_number(self.lookup_out_insert.upper())
        self.lkv_val_Col  = self.column_label_to_number(self.lookup_val_col.upper())

        # Define lookup list
        self.lookup_value = []
        self.lookup_output = []

        for row in self.sheet2.iter_rows(values_only=True):
            self.lookup_value.append(row[self.lkv_val_Col -1])
            
        for value in self.lookup_value:
            for vlook_row in self.sheet1.iter_rows(values_only=True):
                if value == None:
                    continue
                if value == vlook_row[self.lkvTbl_Col - 1]:
                    self.lookup_output.append(vlook_row[self.lkvTbl_Result_Col -1])
    
        (self.head_fill, self.col_fill) =self.color_styling("f2c43d","fce4d6")

        # Insert header with boarder and alignment styling
        self.sheet2.cell(row=2, column= self.lkv_out_insert).value = self.header
        self.sheet2.cell(row=2, column= self.lkv_out_insert).border = self.border
        self.sheet2.cell(row=2, column= self.lkv_out_insert).alignment = self.center_alignment
        self.sheet2.cell(row=2, column= self.lkv_out_insert).fill = self.head_fill

        for i, output in enumerate(self.lookup_output):
            cell = self.sheet2.cell(row=i+3, column= self.lkv_out_insert) 
            cell.value = output
            cell.border = self.border
            cell.alignment = self.center_alignment
            cell.fill =self.col_fill

        self.workbook.save(self.filename)

    def vlookup_Stackup_Table(self):
        #Character to Int Conversion
        self.lkvTbl_Col  = self.column_label_to_number(self.lookup_Tbl_column.upper())
        self.lkvTbl_Result_Col = self.column_label_to_number(self.lookup_Tbl_output.upper())
        self.lkv_out_insert = self.column_label_to_number(self.lookup_out_insert.upper())

        # Define lookup list
        self.lookup_value = ["Layer Name"]
        self.lookup_output = ["Layer Name"]

        self.DDR_14L = ["L1","L14","L3","L5","L10","L12","L1","L14","L3","L5","L10","L12"]
        self.DDR_16L = ["L1","L14","L3","L5","L7","L12","L1","L14","L3","L5","L7","L12",]
        self.DDR_18L = ["L1","L14","L3","L5","L7","L12","L1","L14","L3","L5","L7","L12",]

        # Retrive total layer and store into lookup_value list
        for row in self.sheet1.iter_rows(values_only=True):
            if row[self.lkvTbl_Col -1] == None:
                continue
            else:
                self.lookup_value.append(row[self.lkvTbl_Col -1])
        
        # check the last index of a list 
        if self.lookup_value[-1] == "L14":
            # Perform Layer comparision 
            for value in self.DDR_14L:
                for vlook_row in self.sheet1.iter_rows(values_only=True):
                    if value == vlook_row[self.lkvTbl_Col - 1]:
                        self.lookup_output.append(vlook_row[self.lkvTbl_Result_Col -1])

        # check the last index of a list 
        if self.lookup_value[-1] == "L16":
            # Perform Layer comparision 
            for value in self.DDR_16L:
                for vlook_row in self.sheet1.iter_rows(values_only=True):
                    if value == vlook_row[self.lkvTbl_Col - 1]:
                        self.lookup_output.append(vlook_row[self.lkvTbl_Result_Col -1])

        # check the last index of a list 
        if self.lookup_value[-1] == "L18":
            # Perform Layer comparision 
            for value in self.DDR_18L:
                for vlook_row in self.sheet1.iter_rows(values_only=True):
                    if value == vlook_row[self.lkvTbl_Col - 1]:
                        self.lookup_output.append(vlook_row[self.lkvTbl_Result_Col -1])

        # print(self.lookup_output)

        (self.head_fill, self.col_fill) =self.color_styling("92d050","fce4d6")

        for i, output in enumerate(self.lookup_output):
            cell = self.sheet2.cell(row=i+2, column= self.lkv_out_insert) 
            cell.value = output
            cell.border = self.border
            cell.alignment = self.center_alignment
            cell.fill = self.col_fill

        self.sheet2.cell(row=2,column= self.lkv_out_insert).fill = self.head_fill

        self.workbook.save(self.filename)

    def vlookup_Routing_layer(self):
        #Character to Int Conversion
        self.lkv_NetNm_Col = self.column_label_to_number(self.NetName_Col.upper())
        self.lkv_Tbl_LyrNm_out = self.column_label_to_number(self.lookup_Tbl_LyrNm_out.upper())
        self.lkv_Tbl_col = self.column_label_to_number(self.lookup_Tbl_column.upper())
        self.lkv_lyrNm_out_insert = self.column_label_to_number(self.lookup_lyrNm_out_insert.upper())

        # Define lookup list
        self.lookup_output = []

        # Create empty lists for column 7 and column 8
        LyrStack_values = []
        BO_DQ_values = []
        BO_DQS_values = []

        # Iterate over the columns and append each column to the list
        # Retrive stackup table DQ value start at row 3 and end at row 14 (currently change to max_row=3
        # for debugging purpose )
        for row in self.sheet2.iter_rows(min_row=3, max_row=3, min_col=7, max_col=9):
            LyrStack_values.append(row[0].value)
            BO_DQ_values.append(row[1].value)
            BO_DQS_values.append(row[2].value)


        self.LyrTable_data = [list(values) for values in zip(LyrStack_values, BO_DQ_values, BO_DQS_values)]

        self.table_data = []
        self.repeat_Netname = []
        self.routing_lyr = []


        # Retrive "NetWidth" Table and store into list
        for vlook_row in self.sheet1.iter_rows( values_only=True):
            self.table_data.append(vlook_row)


        #  Loop through Layer Table 
        for data in self.LyrTable_data:
            #Loop through Memory Sheet
            for Netnm_val in self.sheet2.iter_rows(min_row=3, values_only=True):
                # Clear/empty the list to store a new repeated NetName 
                self.repeat_Netname = []
                # Loop Through NetWidth table and store entire row_data into self.repeat_Netname
                for row_data in self.table_data:
                    if Netnm_val[self.lkv_NetNm_Col -1] == row_data[self.lkv_Tbl_col -1]:
                        self.repeat_Netname.append(row_data)
                    else:
                        continue
                
                # Loop through self.repeat_Netname and then check against with 
                # NetName and break the loop if statement are true
                for i in self.repeat_Netname:
                    # Clear/empty the list to store new routing results
                    self.routing_lyr = []
                    if data[0] == i[4]:
                        # print(key, i[1],i[4],i[0])
                        self.routing_lyr.append(i[self.lkv_Tbl_LyrNm_out-1])
                        break
                    else:
                        continue

                # Check if self.routing_lyr list is not empty else append "#N\A"
                if len(self.routing_lyr)>0:
                    self.lookup_output.append(self.routing_lyr[0])
                    self.routing_lyr = []
                    continue
                else:
                    self.lookup_output.append("#N/A")
                    continue
    

        # for i in self.lookup_output:
        #     print(i)
        
        (self.head_fill, self.col_fill) =self.color_styling("f2c43d","fce4d6")

        # Insert header with boarder and alignment styling
        self.sheet2.cell(row=2, column= self.lkv_lyrNm_out_insert).value = self.header_LyrNm
        self.sheet2.cell(row=2, column= self.lkv_lyrNm_out_insert).border = self.border
        self.sheet2.cell(row=2, column= self.lkv_lyrNm_out_insert).alignment = self.center_alignment
        self.sheet2.cell(row=2, column= self.lkv_lyrNm_out_insert).fill = self.head_fill        

        for i, output in enumerate(self.lookup_output):
            cell = self.sheet2.cell(row=i+3, column= self.lkv_lyrNm_out_insert) 
            cell.value = output
            cell.border = self.border
            cell.alignment = self.center_alignment
            cell.fill = self.col_fill

        self.workbook.save(self.filename)

    def vlookup_TotalLength(self):
        #Character to Int Conversion
        self.lkv_NetNm_Col = self.column_label_to_number(self.NetName_Col.upper())
        self.lkv_Tbl_TtlLgth_out = self.column_label_to_number(self.lookup_Tbl_TtlLgth_out.upper())
        self.lkv_Tbl_col = self.column_label_to_number(self.lookup_Tbl_column.upper())
        self.lkv_TtLgth_out_insert = self.column_label_to_number(self.lookup_TtlLgth_out_insert.upper())

        # Define lookup list
        self.lookup_output = []
        self.table_data = []
        self.repeat_Netname = []
        self.Totallength = []

        # Retrive "NetWidth" Table and store into list
        for vlook_row in self.sheet1.iter_rows( values_only=True):
            self.table_data.append(vlook_row)

        #Loop through Memory Sheet
        for Netnm_val in self.sheet2.iter_rows(min_row=3, values_only=True):
            # Clear/empty the list to store a new repeated NetName 
            self.repeat_Netname = [] 
            #Loop Through NetWidth table and store entire row_data into self.repeat_Netname
            for row_data in self.table_data:
                if Netnm_val[self.lkv_NetNm_Col -1] == row_data[self.lkv_Tbl_col -1]:
                    self.repeat_Netname.append(row_data)
                else:
                    continue
            
            # Loop through self.repeat_Netname and then check against with 
            # NetName and break the loop if statement are true
            for i in self.repeat_Netname:
                # Clear/empty the list to store new routing results
                self.Totallength = []
                if Netnm_val[self.lkv_NetNm_Col -1] == i[self.lkv_Tbl_col -1]:
                    self.Totallength.append(i[self.lkv_Tbl_TtlLgth_out-1])
                    break
                else:
                    continue

            # Check if self.Totallength list is not empty else append "#N\A"
            if len(self.Totallength)>0:
                self.lookup_output.append(self.Totallength[0])
                self.Totallength = []
                continue
            else:
                self.lookup_output.append("#N/A")
                continue
    
        # for i in self.lookup_output:
        #     print(i)

        (self.head_fill, self.col_fill) =self.color_styling("f2c43d","fce4d6")
        
        # Insert header with boarder and alignment styling
        self.sheet2.cell(row=2, column= self.lkv_TtLgth_out_insert).value = self.header_TtlLgth
        self.sheet2.cell(row=2, column= self.lkv_TtLgth_out_insert).border = self.border
        self.sheet2.cell(row=2, column= self.lkv_TtLgth_out_insert).alignment = self.center_alignment
        self.sheet2.cell(row=2, column= self.lkv_TtLgth_out_insert).fill = self.head_fill

        for i, output in enumerate(self.lookup_output):
            cell = self.sheet2.cell(row=i+3, column= self.lkv_TtLgth_out_insert) 
            cell.value = output
            cell.border = self.border
            cell.alignment = self.center_alignment
            cell.fill = self.col_fill

        self.workbook.save(self.filename)

    def vlookup_RoutePerMbdg(self):
        #Character to Int Conversion
        self.lkv_Tbl_col = self.column_label_to_number(self.lookup_Tbl_column.upper())
        self.lkv_val_col = self.column_label_to_number(self.lookup_val_col.upper())
        self.lkv_out_insert = self.column_label_to_number(self.lookup_out_insert.upper())

        # Define lookup list
        self.lookup_output = []

        #  Loop through Layer Table 
        for layerNm in self.sheet1.iter_rows(min_row=3, max_row=3, values_only=True):
            #Loop through Memory Sheet
            for RouteLyr in self.sheet2.iter_rows(min_row=3, values_only=True):
                if layerNm[self.lkv_Tbl_col-1] == RouteLyr[self.lkv_val_col-1]:
                    self.lookup_output.append("OK")
                else:
                    self.lookup_output.append("NO")
    

        # for i in self.lookup_output:
        #     print(i)
        
        (self.head_fill, self.col_fill_Green) =self.color_styling("f2c43d","c6efce")
        (self.head_fill, self.col_fill_Red) =self.color_styling("f2c43d","ff8172")

        # Insert header with boarder and alignment styling
        self.sheet2.cell(row=2, column= self.lkv_out_insert).value = self.header
        self.sheet2.cell(row=2, column= self.lkv_out_insert).border = self.border
        self.sheet2.cell(row=2, column= self.lkv_out_insert).alignment = self.center_alignment
        self.sheet2.cell(row=2, column= self.lkv_out_insert).fill = self.head_fill        

        for i, output in enumerate(self.lookup_output):
            if output == "OK":
                cell = self.sheet2.cell(row=i+3, column= self.lkv_out_insert) 
                cell.value = output
                cell.border = self.border
                cell.alignment = self.center_alignment
                cell.fill = self.col_fill_Green
            else:
                cell = self.sheet2.cell(row=i+3, column= self.lkv_out_insert) 
                cell.value = output
                cell.border = self.border
                cell.alignment = self.center_alignment
                cell.fill = self.col_fill_Red


        self.workbook.save(self.filename)

    def vlookup_BreakoutLength(self):
        #Character to Int Conversion
        self.lkv_Tbl_col = self.column_label_to_number(self.lookup_Tbl_column.upper())
        self.lkv_Tbl_linWdth_Col = self.column_label_to_number(self.lookup_Tbl_col_LineWdt.upper())
        self.lkv_Tbl_out = self.column_label_to_number(self.lookup_Tbl_output.upper())
        self.lkv_val_col = self.column_label_to_number(self.lookup_val_col.upper())
        self.lkv_out_insert_Col = self.column_label_to_number(self.lookup_out_insert.upper())

        # Define lookup list
        self.lookup_output = []

        # Create empty lists for column 7 and column 8
        LyrStack_values = []
        BO_DQ_values = []
        BO_DQS_values = []
        Bus_DQ_values = []
        Bus_DQS_values = []

        # Iterate over the columns and append each column to the list
        # Retrive stackup table DQ value start at row 3 and end at row 14 (currently change to max_row=3
        # for debugging purpose )
        for row in self.sheet2.iter_rows(min_row=3, max_row=3, min_col=7, max_col=11):
            LyrStack_values.append(row[0].value)
            BO_DQ_values.append(row[1].value)
            BO_DQS_values.append(row[2].value)
            Bus_DQ_values.append(row[3].value)
            Bus_DQS_values.append(row[4].value)


        self.LyrTable_data = [list(values) for values in zip(LyrStack_values, 
                                                            BO_DQ_values, 
                                                            BO_DQS_values,
                                                            Bus_DQ_values,
                                                            Bus_DQS_values)]

        self.table_data = []
        self.repeat_Netname = []
        self.BO_Length = []

        # Retrive "NetWidth" Table and store into list
        for vlook_row in self.sheet1.iter_rows( values_only=True):
            self.table_data.append(vlook_row)

        #  Loop through Layer Table 
        for data in self.LyrTable_data:
            #Loop through Memory Sheet
            for Netnm_val in self.sheet2.iter_rows(min_row=3, values_only=True):
                # Clear/empty the list to store a new repeated NetName 
                self.repeat_Netname = []
                # Loop Through NetWidth table and store entire row_data into self.repeat_Netname
                for row_data in self.table_data:
                    if Netnm_val[self.lkv_val_col -1] == row_data[self.lkv_Tbl_col -1]:
                        self.repeat_Netname.append(row_data)
                    else:
                        continue
                
                # Loop through self.repeat_Netname and then check against with 
                # layer name and DQ/DQS value and break the loop if data or value of
                # (DQ/DQS) statement are true
                # NOTE: e.g of data's list on channel A => ("TOP","3.66","3.66","4.3","5.0")
                for i in self.repeat_Netname:
                     # Clear/empty the list to store new 
                    self.BO_Length = []
                    if "DQS" in i[0]:
                        if data[2] == i[self.lkv_Tbl_linWdth_Col-1]:
                            # print(value, i[1],i[0])
                            self.BO_Length.append(i[self.lkv_Tbl_out-1])
                            break
                        elif data[4] == i[self.lkv_Tbl_linWdth_Col-1]:
                            self.BO_Length.append(i[self.lkv_Tbl_out-1])
                            break
                        else:
                            continue
                    else:
                        if data[1] == i[self.lkv_Tbl_linWdth_Col-1]:
                            # print(value, i[1],i[0])
                            self.BO_Length.append(i[self.lkv_Tbl_out-1])
                            break
                        elif data[3] == i[self.lkv_Tbl_linWdth_Col-1]:
                            self.BO_Length.append(i[self.lkv_Tbl_out-1])
                            break
                        else:
                            continue

                # Check if self.BO_Length list is not empty else append "#N\A"
                if len(self.BO_Length)>0:
                    self.lookup_output.append(self.BO_Length[0])
                    self.BO_Length = []
                    continue
                else:
                    self.lookup_output.append("#N/A")
                    continue
    

        # for i in self.lookup_output:
        #     print(i)
        
        (self.head_fill, self.col_fill) =self.color_styling("f2c43d","fce4d6")

        # Insert header with boarder and alignment styling
        self.sheet2.cell(row=2, column= self.lkv_out_insert_Col).value = self.header
        self.sheet2.cell(row=2, column= self.lkv_out_insert_Col).border = self.border
        self.sheet2.cell(row=2, column= self.lkv_out_insert_Col).alignment = self.center_alignment
        self.sheet2.cell(row=2, column= self.lkv_out_insert_Col).fill = self.head_fill        

        for i, output in enumerate(self.lookup_output):
            cell = self.sheet2.cell(row=i+3, column= self.lkv_out_insert_Col) 
            cell.value = output
            cell.border = self.border
            cell.alignment = self.center_alignment
            cell.fill = self.col_fill

        self.workbook.save(self.filename)

    def vlookup_impedance(self):
        #Character to Int Conversion
        self.lkv_Tbl_col = self.column_label_to_number(self.lookup_Tbl_column.upper())
        self.lkv_val_col = self.column_label_to_number(self.lookup_val_col.upper())
        self.lkv_out_insert = self.column_label_to_number(self.lookup_out_insert.upper())
        self.lkv_NetNm_Col = self.column_label_to_number(self.NetName_Col.upper())

        # Define lookup list
        self.lookup_output = []

        #  Loop through Layer Table 
        for DDRLength in self.sheet1.iter_rows(min_row=17, max_row=17, values_only=True):
            #Loop through Memory Sheet
            for TotalLength in self.sheet2.iter_rows(min_row=3, values_only=True):
                # print(TotalLength[self.lkv_NetNm_Col-1] ,"no")
                if TotalLength[self.lkv_NetNm_Col-1] is None:
                    self.lookup_output.append("#N/A")
                elif "DQS" in TotalLength[self.lkv_NetNm_Col-1]:
                    self.lookup_output.append("#N/A")
                elif int(DDRLength[self.lkv_Tbl_col-1]) >= int(TotalLength[self.lkv_val_col-1]):
                    self.lookup_output.append("OK")
                else:
                    self.lookup_output.append("40OHM")
    

        # for i in self.lookup_output:
        #     print(i)
        
        (self.head_fill, self.col_fill_Green) =self.color_styling("f2c43d","c6efce")
        (self.head_fill, self.col_fill_Red) =self.color_styling("f2c43d","ff8172")
        (self.head_fill, self.col_fill_Grey) =self.color_styling("f2c43d","d0cece")

        # Insert header with boarder and alignment styling
        self.sheet2.cell(row=2, column= self.lkv_out_insert).value = self.header
        self.sheet2.cell(row=2, column= self.lkv_out_insert).border = self.border
        self.sheet2.cell(row=2, column= self.lkv_out_insert).alignment = self.center_alignment
        self.sheet2.cell(row=2, column= self.lkv_out_insert).fill = self.head_fill        

        for i, output in enumerate(self.lookup_output):
            if output == "OK":
                cell = self.sheet2.cell(row=i+3, column= self.lkv_out_insert) 
                cell.value = output
                cell.border = self.border
                cell.alignment = self.center_alignment
                cell.fill = self.col_fill_Green
            elif output == "#N/A":
                cell = self.sheet2.cell(row=i+3, column= self.lkv_out_insert) 
                cell.value = output
                cell.border = self.border
                cell.alignment = self.center_alignment
                cell.fill = self.col_fill_Grey             
            else:
                cell = self.sheet2.cell(row=i+3, column= self.lkv_out_insert) 
                cell.value = output
                cell.border = self.border
                cell.alignment = self.center_alignment
                cell.fill = self.col_fill_Red


        self.workbook.save(self.filename)