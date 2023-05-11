# import pandas as pd

# # Create the first DataFrame
# df1 = pd.DataFrame({'key': ['A', 'B', 'C', 'D'], 'value': [1, 2, 3, 4]})

# # Create the second DataFrame
# df2 = pd.DataFrame({'key': ['B', 'D'], 'value2': [5, 6]})

# # Perform a lookup operation using map()
# df1['value2'] = df1['key'].map(df2.set_index('key')['value2'])

# print(df1['value2'])
#-------------------------------------------------------------------#
# import openpyxl

# # Open the Excel file
# workbook = openpyxl.load_workbook('2023-05-10 172439.xlsx')

# # Select the sheet to work with
# sheet = workbook['Layer_Stackup']

# # Define the cell to look up
# lookup_value_L1 = sheet['A30'].value

# # Define the range to search in
# table_array = sheet['A1:E32']

# # Perform the VLOOKUP operation
# result_L1 = None
# result_L14 = None
# result_L3 = None
# result_L5 = None
# result_L10 = None
# result_L12 = None

# for row in table_array:
#     # print(row[0].value)
#     if row[0].value == "L1":
#         result_L1 = row[1].value
#     if row[0].value == "L14":
#         result_L14 = row[1].value
#     if row[0].value == "L3":
#         result_L3 = row[1].value
#     if row[0].value == "L5":
#         result_L5 = row[1].value
#     if row[0].value == "L10":
#         result_L10 = row[1].value
#     if row[0].value == "L12":
#         result_L12 = row[1].value


# # Print the result
# print(result_L1)
# print(result_L14)
# print(result_L3)
# print(result_L5)
# print(result_L10)
# print(result_L12)

#---------------------------------------------------------------#
import openpyxl

# Open the Excel file
workbook = openpyxl.load_workbook('2023-05-10 172439.xlsx')

# Select the sheet to work with
sheet1 = workbook['Pin_Net']
sheet2 = workbook['MEMORY']

# Define lookup list
lookup_value = []
lookup_output = []

def column_label_to_number(label):
    total = 0
    for i, char in enumerate(reversed(label)):
        value = ord(char) - 64
        total += value * (26 ** i)
    return total

lkvTbl_Col  = column_label_to_number("C")
lkvTbl_Result_Col = column_label_to_number("D")
lkv_value_Col  = column_label_to_number("K")

# # # Define the column index to extract data from
# # col_idx = 0  # column B

for row in sheet2.iter_rows(values_only=True):
    # print(row[10])
    lookup_value.append(row[lkv_value_Col -1])
    
for value in lookup_value:
    for vlook_row in sheet1.iter_rows(values_only=True):
        if value == None:
            continue
        if value == vlook_row[lkvTbl_Col - 1]:
            lookup_output.append(vlook_row[lkvTbl_Result_Col -1])

# print(len(lookup_output))
for i in lookup_output:
    print(i)

# # for i in lookup_value:
# #     print(i)

# #Required Input
# #1)sheetname 1
# #2)sheetname 2
# #3)sheet2 lookup value column
# #4)sheet1 lookup column
# #5)sheet1 lookup output
# #6)shee2 insert col